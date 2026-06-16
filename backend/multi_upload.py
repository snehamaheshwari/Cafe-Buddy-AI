"""
Multi-dataset upload routes for Cafe Buddy.
Handles three distinct data types:
  1. Financial Data   — costs, margins, revenue, commissions
  2. POS Billing Data — orders, items, payments, timestamps
  3. Customer Data    — CRM, loyalty, preferences, feedback
Supports both .xlsx / .xls (Excel) and .csv files.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

import pandas as pd
from fastapi import APIRouter, File, HTTPException, Query, Request, UploadFile

import data_store

router = APIRouter()


def _get_tenant_id(request: Optional[Request]) -> str:
    """Extract tenant_id from JWT Bearer token; fall back to system tenant."""
    import auth_utils as _au
    import tenant_store as _ts
    if not request:
        return _ts.SYSTEM_TENANT_ID
    auth_header = request.headers.get("Authorization", "")
    return _au.extract_tenant_id(auth_header) or _ts.SYSTEM_TENANT_ID

# ─────────────────────────────────────────────
# SHARED COLUMN-DETECTION UTILITY
# ─────────────────────────────────────────────

import re as _re

def _norm(s: str) -> str:
    """Normalise a column name: strip non-printable chars, collapse whitespace, lowercase."""
    s = _re.sub(r'[^\x20-\x7E]', ' ', str(s))   # strip non-ASCII printable
    s = _re.sub(r'\s+', ' ', s).strip().lower()
    return s


def _detect(df_cols: list, aliases: dict, key: str) -> Optional[str]:
    """
    4-pass detection:
      1. Exact normalised match
      2. Any slash/pipe fragment matches an alias
      3. Alias is a prefix or suffix of the normalised column name
      4. Column name contains the alias as a substring
    """
    norm_map = {_norm(c): c for c in df_cols}

    for alias in aliases.get(key, []):
        al = _norm(alias)
        if al in norm_map:
            return norm_map[al]

    for col in df_cols:
        parts = [p.strip().lower() for p in _norm(col).replace("|", "/").split("/")]
        for alias in aliases.get(key, []):
            al = _norm(alias)
            if al in parts:
                return col

    for col in df_cols:
        cl = _norm(col)
        for alias in aliases.get(key, []):
            al = _norm(alias)
            if cl.startswith(al) or cl.endswith(al):
                return col

    for col in df_cols:
        cl = _norm(col)
        for alias in aliases.get(key, []):
            al = _norm(alias)
            if len(al) >= 4 and al in cl:
                return col

    return None


def _read_file(raw: bytes, filename: str) -> pd.DataFrame:
    """Auto-detect CSV vs Excel."""
    if filename.lower().endswith(".csv"):
        return pd.read_csv(io.BytesIO(raw))
    return pd.read_excel(io.BytesIO(raw), engine="openpyxl")


def _safe_float(val, default: float = 0.0) -> float:
    try:
        if pd.isna(val):
            return default
        s = str(val).replace(",", "").replace("₹", "").replace("%", "").strip()
        return float(s)
    except Exception:
        return default


def _safe_str(val, default: str = "") -> str:
    try:
        if pd.isna(val):
            return default
        return str(val).strip()
    except Exception:
        return default


# ─────────────────────────────────────────────
# 1. FINANCIAL DATA
# ─────────────────────────────────────────────

FIN_ALIASES = {
    "date": [
        "date", "day", "period", "week", "month", "order date", "sale date",
        "transaction date", "report date",
    ],
    "daily_revenue": [
        "daily revenue", "revenue", "total revenue", "gross revenue",
        "net revenue", "sales", "turnover", "gross sales", "income",
    ],
    "gross_margin": [
        "gross margin", "gross margin %", "gross margin pct", "gm",
        "gm %", "gross profit %", "gross profit pct",
    ],
    "net_profit": [
        "net profit", "net profit %", "net income", "profit", "nett profit",
        "net margin", "bottom line",
    ],
    "food_cost": [
        "food cost", "food cost %", "food cost percentage", "cogs %",
        "cost of goods sold %", "material cost %", "ingredient cost",
    ],
    "labor_cost": [
        "labor cost", "labour cost", "labor cost %", "labour cost %",
        "staff cost", "salary cost", "manpower cost", "payroll",
    ],
    "electricity": [
        "electricity", "electricity cost", "power cost", "utility",
        "utilities", "electricity bill", "power bill",
    ],
    "rent": [
        "rent", "rental", "lease", "premises cost", "shop rent",
        "store rent", "rent cost",
    ],
    "marketing": [
        "marketing", "marketing spend", "advertising", "ads",
        "promotion cost", "marketing cost", "ad spend",
    ],
    "packaging": [
        "packaging", "packaging cost", "packing cost",
        "packaging material", "packing material",
    ],
    "commission": [
        "commission", "swiggy commission", "zomato commission",
        "platform commission", "aggregator commission",
        "delivery commission", "online commission",
    ],
}


def _parse_financial(df: pd.DataFrame, filename: str) -> tuple[list, dict]:
    df.columns = [str(c).strip() for c in df.columns]

    date_col    = _detect(df.columns, FIN_ALIASES, "date")
    rev_col     = _detect(df.columns, FIN_ALIASES, "daily_revenue")
    gm_col      = _detect(df.columns, FIN_ALIASES, "gross_margin")
    np_col      = _detect(df.columns, FIN_ALIASES, "net_profit")
    fc_col      = _detect(df.columns, FIN_ALIASES, "food_cost")
    lc_col      = _detect(df.columns, FIN_ALIASES, "labor_cost")
    el_col      = _detect(df.columns, FIN_ALIASES, "electricity")
    rent_col    = _detect(df.columns, FIN_ALIASES, "rent")
    mkt_col     = _detect(df.columns, FIN_ALIASES, "marketing")
    pkg_col     = _detect(df.columns, FIN_ALIASES, "packaging")
    comm_col    = _detect(df.columns, FIN_ALIASES, "commission")

    if not date_col:
        raise ValueError("Could not find a 'Date' column in the financial file.")
    if not rev_col:
        raise ValueError("Could not find a 'Revenue / Daily Revenue' column.")

    records, skipped = [], 0
    for _, row in df.iterrows():
        try:
            raw_date = row[date_col]
            if pd.isna(raw_date):
                skipped += 1; continue
            date_str = pd.to_datetime(raw_date).strftime("%Y-%m-%d")

            revenue = _safe_float(row[rev_col]) if rev_col else 0.0
            if revenue <= 0:
                skipped += 1; continue

            records.append({
                "date":            date_str,
                "daily_revenue":   revenue,
                "gross_margin_pct": _safe_float(row[gm_col])  if gm_col  else None,
                "net_profit":      _safe_float(row[np_col])   if np_col  else None,
                "food_cost_pct":   _safe_float(row[fc_col])   if fc_col  else None,
                "labor_cost_pct":  _safe_float(row[lc_col])   if lc_col  else None,
                "electricity":     _safe_float(row[el_col])   if el_col  else None,
                "rent":            _safe_float(row[rent_col]) if rent_col else None,
                "marketing":       _safe_float(row[mkt_col])  if mkt_col  else None,
                "packaging":       _safe_float(row[pkg_col])  if pkg_col  else None,
                "commission":      _safe_float(row[comm_col]) if comm_col else None,
            })
        except Exception:
            skipped += 1

    # Auto-fix: Excel sometimes stores percentage columns as decimals (0.32 → 32%).
    # Fix is applied PER ROW: if a value is in (0, 1.5] it is almost certainly
    # fractional encoding (e.g. 0.32 means 32%) and is multiplied by 100.
    # Values > 1.5 are treated as already in percentage form (e.g. 65.0 = 65%).
    # This handles mixed-format files where some rows have 65 and others have 0.65.
    _pct_fields = ["gross_margin_pct", "food_cost_pct", "labor_cost_pct", "net_profit"]
    for field in _pct_fields:
        for r in records:
            val = r.get(field)
            if val is not None and 0 < val <= 1.5:
                r[field] = round(val * 100, 2)

    dates = sorted(set(r["date"] for r in records))
    detected = {k: v for k, v in {
        "date": date_col, "daily_revenue": rev_col, "gross_margin": gm_col,
        "net_profit": np_col, "food_cost": fc_col, "labor_cost": lc_col,
        "electricity": el_col, "rent": rent_col, "marketing": mkt_col,
        "packaging": pkg_col, "commission": comm_col,
    }.items() if v}

    info = {
        "filename": filename, "rows": len(records), "skipped": skipped,
        "columns_detected": detected,
        "date_range": {"from": dates[0], "to": dates[-1]} if dates else {},
        "unique_dates": len(dates),
        "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    return records, info


# ─────────────────────────────────────────────
# 2. POS BILLING DATA
# ─────────────────────────────────────────────

POS_ALIASES = {
    "order_id": [
        "order id", "order_id", "bill id", "bill no", "invoice id",
        "transaction id", "receipt no", "order number", "bill number",
    ],
    "timestamp": [
        "timestamp", "order time", "order timestamp", "date time",
        "datetime", "order date", "date", "created at",
    ],
    # "revenue" is the primary money column — exact match wins over bill_amount
    "revenue": [
        "revenue", "net revenue", "net sales", "sale amount",
        "total revenue", "gross revenue",
    ],
    # bill_amount is a fallback if there is no dedicated revenue column;
    # NOTE: "amount" deliberately excluded to avoid matching "GST Amount"
    "bill_amount": [
        "bill amount", "total amount", "order amount", "invoice amount",
        "grand total", "final amount", "payable amount", "total",
    ],
    "price": [
        "price", "rate", "unit price", "selling price", "mrp",
        "item price", "price / rate",
    ],
    "cost": [
        "cost", "cogs", "cost of goods", "item cost", "unit cost",
        "food cost amount", "ingredient cost",
    ],
    "category": [
        "category", "item category", "product category", "type",
        "item type", "menu category", "food category",
    ],
    "gst": [
        "gst", "gst amount", "tax", "taxes", "vat", "tax amount",
        "sgst", "cgst",
    ],
    "discount": [
        "discount", "discount amount", "offer", "discount value",
        "promo discount", "coupon discount",
    ],
    "coupon": [
        "coupon", "coupon code", "promo code", "voucher", "coupon used",
        "offer code", "discount code",
    ],
    "item_name": [
        "item name", "item name / product", "product", "item",
        "menu item", "food item", "dish",
    ],
    "quantity": [
        "quantity", "qty", "count", "units", "item quantity",
        "quantity / qty",
    ],
    "payment_mode": [
        "payment mode", "payment method", "pay mode", "mode of payment",
        "payment type", "paid via", "pay via",
    ],
    "platform": [
        "platform", "delivery platform", "channel", "order source",
        "platform / channel", "delivery channel", "order channel",
        "source",
    ],
    "hour": [
        "hour", "time", "order time", "time slot", "peak hour",
        "time period", "rush hour", "meal period", "slot",
    ],
    "daypart": [
        "daypart", "day part", "meal period", "time of day", "session",
        "shift",
    ],
    "customer_phone": [
        "customer phone", "customer mobile", "customer contact",
        "phone", "mobile", "contact number",
    ],
    "cafe_location": [
        "cafe location", "location", "branch", "outlet", "store",
        "cafe branch", "restaurant location",
    ],
    "repeat_customer": [
        "repeat customer", "returning customer", "loyal customer",
        "new customer", "is repeat", "customer type",
    ],
    "cancel_reason": [
        "cancellation reason", "cancel reason", "cancelled reason",
        "why cancelled", "cancellation",
    ],
    "refund_reason": [
        "refund reason", "return reason", "why refunded",
    ],
}


def _parse_hour(val) -> int:
    """Extract integer hour from various formats: datetime, 'HH:MM', int, float."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 12
    if isinstance(val, (int, float)):
        return int(val) % 24
    s = str(val).strip()
    # Handle "HH:MM" or "HH:MM:SS" string from Excel Hour column
    if ":" in s:
        try:
            return int(s.split(":")[0]) % 24
        except Exception:
            pass
    # Try parsing as a full datetime string
    try:
        return pd.to_datetime(s).hour
    except Exception:
        return 12


def _parse_pos(df: pd.DataFrame, filename: str) -> tuple[list, dict]:
    df.columns = [str(c).strip() for c in df.columns]

    order_col  = _detect(df.columns, POS_ALIASES, "order_id")
    ts_col     = _detect(df.columns, POS_ALIASES, "timestamp")
    rev_col    = _detect(df.columns, POS_ALIASES, "revenue")
    bill_col   = _detect(df.columns, POS_ALIASES, "bill_amount")
    price_col  = _detect(df.columns, POS_ALIASES, "price")
    cost_col   = _detect(df.columns, POS_ALIASES, "cost")
    cat_col    = _detect(df.columns, POS_ALIASES, "category")
    gst_col    = _detect(df.columns, POS_ALIASES, "gst")
    disc_col   = _detect(df.columns, POS_ALIASES, "discount")
    coup_col   = _detect(df.columns, POS_ALIASES, "coupon")
    item_col   = _detect(df.columns, POS_ALIASES, "item_name")
    qty_col    = _detect(df.columns, POS_ALIASES, "quantity")
    pay_col    = _detect(df.columns, POS_ALIASES, "payment_mode")
    plat_col   = _detect(df.columns, POS_ALIASES, "platform")
    hour_col   = _detect(df.columns, POS_ALIASES, "hour")
    daypart_col= _detect(df.columns, POS_ALIASES, "daypart")
    phone_col  = _detect(df.columns, POS_ALIASES, "customer_phone")
    loc_col    = _detect(df.columns, POS_ALIASES, "cafe_location")
    repeat_col = _detect(df.columns, POS_ALIASES, "repeat_customer")
    cancel_col = _detect(df.columns, POS_ALIASES, "cancel_reason")
    refund_col = _detect(df.columns, POS_ALIASES, "refund_reason")

    # Revenue column takes priority; bill_amount is a fallback
    amount_col = rev_col or bill_col

    if not ts_col and not order_col:
        raise ValueError("Could not find an 'Order ID' or 'Date' column.")
    if not amount_col:
        raise ValueError("Could not find a 'Revenue' / 'Bill Amount' column.")

    records, skipped = [], 0
    for _, row in df.iterrows():
        try:
            # ── Date ──────────────────────────────────────────────────
            raw_ts = row[ts_col] if ts_col else None
            if raw_ts is not None and not pd.isna(raw_ts):
                date_s = pd.to_datetime(raw_ts).strftime("%Y-%m-%d")
            else:
                date_s = datetime.now().strftime("%Y-%m-%d")

            # ── Hour ──────────────────────────────────────────────────
            # Prefer a dedicated Hour column; fall back to parsing timestamp
            if hour_col:
                hour = _parse_hour(row[hour_col])
            elif ts_col and raw_ts is not None and not pd.isna(raw_ts):
                hour = pd.to_datetime(raw_ts).hour
            else:
                hour = 12

            # ── Core financials (exact values from Excel) ─────────────
            revenue  = _safe_float(row[amount_col])
            if revenue <= 0:
                skipped += 1; continue

            qty      = _safe_float(row[qty_col], 1.0)       if qty_col    else 1.0
            # Use exact Price column if present; compute only as fallback
            price    = _safe_float(row[price_col])           if price_col  else round(revenue / max(qty, 1), 2)
            # Use exact Cost column if present; compute only as fallback
            cost     = _safe_float(row[cost_col])            if cost_col   else round(revenue * 0.38, 2)

            item     = _safe_str(row[item_col], "Unknown")   if item_col   else "Unknown"
            platform = _safe_str(row[plat_col], "Dine-in")   if plat_col   else "Dine-in"
            category = _safe_str(row[cat_col], "POS Order")  if cat_col    else "POS Order"

            records.append({
                # analytics-compatible keys (used by get_data() → AI layers)
                "date":             date_s,
                "item_name":        item,
                "category":         category,
                "quantity":         qty,
                "price":            price,
                "revenue":          revenue,
                "cost":             cost,
                "platform":         platform,
                # POS-specific keys
                "order_id":         _safe_str(row[order_col])   if order_col   else "",
                "bill_amount":      revenue,
                "gst":              _safe_float(row[gst_col])   if gst_col     else None,
                "discount":         _safe_float(row[disc_col])  if disc_col    else None,
                "coupon":           _safe_str(row[coup_col])    if coup_col    else "",
                "payment_mode":     _safe_str(row[pay_col])     if pay_col     else "",
                "hour":             hour,
                "peak_hour":        _safe_str(row[daypart_col]) if daypart_col else _hour_label(hour),
                "customer_phone":   _safe_str(row[phone_col])   if phone_col   else "",
                "cafe_location":    _safe_str(row[loc_col])     if loc_col     else "",
                "repeat":           _safe_str(row[repeat_col])  if repeat_col  else "",
                "cancel_reason":    _safe_str(row[cancel_col])  if cancel_col  else "",
                "refund_reason":    _safe_str(row[refund_col])  if refund_col  else "",
            })
        except Exception:
            skipped += 1

    dates   = sorted(set(r["date"] for r in records))
    detected = {k: v for k, v in {
        "order_id": order_col, "timestamp": ts_col,
        "revenue": rev_col, "bill_amount": bill_col,
        "price": price_col, "cost": cost_col, "category": cat_col,
        "gst": gst_col, "discount": disc_col, "coupon": coup_col,
        "item_name": item_col, "quantity": qty_col, "payment_mode": pay_col,
        "platform": plat_col, "hour": hour_col, "daypart": daypart_col,
        "customer_phone": phone_col, "cafe_location": loc_col,
        "repeat_customer": repeat_col, "cancel_reason": cancel_col,
        "refund_reason": refund_col,
    }.items() if v}

    info = {
        "filename": filename, "rows": len(records), "skipped": skipped,
        "columns_detected": detected,
        "date_range": {"from": dates[0], "to": dates[-1]} if dates else {},
        "unique_dates": len(dates),
        "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    return records, info


def _hour_label(h: int) -> str:
    if 6  <= h < 11: return "Breakfast (6–11)"
    if 11 <= h < 15: return "Lunch (11–15)"
    if 15 <= h < 18: return "Afternoon (15–18)"
    if 18 <= h < 22: return "Dinner (18–22)"
    return "Late Night (22+)"


# ─────────────────────────────────────────────
# 3. CUSTOMER DATA
# ─────────────────────────────────────────────

CUST_ALIASES = {
    "name": [
        "customer name", "name", "full name", "client name",
        "guest name", "user name", "customer",
    ],
    "phone": [
        "phone / contact", "phone/contact", "phone",
        "phone number", "mobile", "mobile number",
        "contact", "contact number", "cell", "whatsapp",
    ],
    "birthday": [
        "birthday / dob", "birthday/dob", "birthday",
        "dob", "date of birth", "birth date", "birth day", "bday",
    ],
    "visit_frequency": [
        "visit frequency", "visits", "frequency", "visit count",
        "total visits", "no of visits", "number of visits",
    ],
    "favorite_items": [
        "favourite items", "favorite items", "preferred items",
        "top items", "fav items", "fav dish", "preferred dish",
    ],
    "avg_order_value": [
        "avg order value", "average order value", "aov",
        "avg order", "avg bill", "average bill", "average spend",
    ],
    "feedback": [
        "feedback / rating", "feedback/rating",
        "feedback", "sentiment", "rating", "review", "score",
        "feedback sentiment", "customer rating", "nps",
    ],
    "preferred_time": [
        "preferred visit time", "preferred time", "ordering time",
        "preferred ordering time", "visit time", "preferred slot", "usual time",
    ],
    "platform_source": [
        "platform source", "platform", "source", "acquisition source",
        "channel", "how did they find us", "referral",
    ],
    "loyalty_points": [
        "loyalty points", "points", "rewards points", "reward points",
        "loyalty", "loyalty score",
    ],
    "gender": ["gender", "sex"],
    "age_group": [
        "age group", "age", "age bracket", "age range",
        "age category",
    ],
}


def _parse_customer(df: pd.DataFrame, filename: str) -> tuple[list, dict]:
    # Strip non-printable chars from column names (handles BOM, zero-width spaces, etc.)
    df.columns = [_re.sub(r'[^\x20-\x7E]', ' ', str(c)).strip() for c in df.columns]

    # Auto-detect header row — some files have a merged title row as pandas header.
    # We scan the first 5 rows (including what pandas read as df.columns) for the
    # row that best matches known column aliases, then use it as the real header.
    all_known = set(_norm(alias) for aliases in CUST_ALIASES.values() for alias in aliases)
    best_score, header_row = 0, -1

    # Check current df.columns first (row index -1 meaning "already the header")
    col_score = sum(1 for c in df.columns if _norm(str(c)) in all_known)
    if col_score >= 3:
        best_score, header_row = col_score, -1  # columns are already correct

    # Check first 5 data rows
    for i in range(min(5, len(df))):
        row_vals = [_norm(str(v)) for v in df.iloc[i] if pd.notna(v)]
        score = sum(1 for v in row_vals if v in all_known)
        if score > best_score:
            best_score, header_row = score, i

    # If a data row scored better than the current columns, promote it to header
    if header_row >= 0:
        df.columns = [_re.sub(r'[^\x20-\x7E]', ' ', str(v)).strip() for v in df.iloc[header_row]]
        df = df.iloc[header_row + 1:].reset_index(drop=True)

    cols = df.columns.tolist()
    name_col   = _detect(cols, CUST_ALIASES, "name")
    phone_col  = _detect(cols, CUST_ALIASES, "phone")
    bday_col   = _detect(cols, CUST_ALIASES, "birthday")
    freq_col   = _detect(cols, CUST_ALIASES, "visit_frequency")
    fav_col    = _detect(cols, CUST_ALIASES, "favorite_items")
    aov_col    = _detect(cols, CUST_ALIASES, "avg_order_value")
    fb_col     = _detect(cols, CUST_ALIASES, "feedback")
    time_col   = _detect(cols, CUST_ALIASES, "preferred_time")
    src_col    = _detect(cols, CUST_ALIASES, "platform_source")
    pts_col    = _detect(cols, CUST_ALIASES, "loyalty_points")
    gender_col = _detect(cols, CUST_ALIASES, "gender")
    age_col    = _detect(cols, CUST_ALIASES, "age_group")

    if not name_col and not phone_col:
        cols_found = list(df.columns)[:8]
        raise ValueError(
            f"Could not find a customer identifier column. "
            f"Expected 'Customer Name' or 'Phone / Contact'. "
            f"Columns detected in your file: {cols_found}. "
            f"Please ensure your file has at least a Name or Phone column."
        )

    records, skipped = [], 0
    for _, row in df.iterrows():
        try:
            name  = _safe_str(row[name_col])  if name_col  else ""
            phone = _safe_str(row[phone_col]) if phone_col else ""
            if not name and not phone:
                skipped += 1; continue

            records.append({
                "name":             name,
                "phone":            phone,
                "birthday":         _safe_str(row[bday_col])  if bday_col  else "",
                "visit_frequency":  _safe_str(row[freq_col])  if freq_col  else "",
                "favorite_items":   _safe_str(row[fav_col])   if fav_col   else "",
                "avg_order_value":  _safe_float(row[aov_col]) if aov_col   else None,
                "feedback":         _safe_str(row[fb_col])    if fb_col    else "",
                "preferred_time":   _safe_str(row[time_col])  if time_col  else "",
                "platform_source":  _safe_str(row[src_col])   if src_col   else "",
                "loyalty_points":   _safe_float(row[pts_col]) if pts_col   else 0.0,
                "gender":           _safe_str(row[gender_col])if gender_col else "",
                "age_group":        _safe_str(row[age_col])   if age_col   else "",
            })
        except Exception:
            skipped += 1

    detected = {k: v for k, v in {
        "name": name_col, "phone": phone_col, "birthday": bday_col,
        "visit_frequency": freq_col, "favorite_items": fav_col,
        "avg_order_value": aov_col, "feedback": fb_col,
        "preferred_time": time_col, "platform_source": src_col,
        "loyalty_points": pts_col, "gender": gender_col, "age_group": age_col,
    }.items() if v}

    info = {
        "filename": filename, "rows": len(records), "skipped": skipped,
        "columns_detected": detected,
        "total_customers": len(records),
        "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    return records, info


# ─────────────────────────────────────────────
# API ENDPOINTS
# ─────────────────────────────────────────────

ALLOWED_TYPES = (".xlsx", ".xls", ".csv")


def _check_ext(filename: str):
    if not any(filename.lower().endswith(ext) for ext in ALLOWED_TYPES):
        raise HTTPException(status_code=400,
                            detail="Only .xlsx, .xls, or .csv files are supported.")


# ── Financial ──────────────────────────────────────────────────────────────────

@router.post("/upload/financial")
async def upload_financial(file: UploadFile = File(...), request: Request = None):
    _check_ext(file.filename)
    raw = await file.read()
    try:
        df = _read_file(raw, file.filename)
        records, info = _parse_financial(df, file.filename)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Parse error: {e}")

    if not records:
        raise HTTPException(status_code=422, detail="No valid rows found.")

    tid = _get_tenant_id(request)
    data_store.set_financial_for_tenant(tid, records, info)
    return {"success": True, "message": f"Loaded {len(records)} financial records.", "info": info}


@router.delete("/upload/financial/clear")
def clear_financial(request: Request = None):
    tid = _get_tenant_id(request)
    data_store.clear_financial_for_tenant(tid)
    return {"success": True}


@router.get("/data/financial/summary")
def financial_summary(request: Request = None):
    tid = _get_tenant_id(request)
    data, info = data_store.get_financial_for_tenant(tid)
    if not data:
        return {"uploaded": False, "info": {}, "summary": {}}

    total_rev  = sum(r["daily_revenue"] for r in data)
    avg_gm     = _avg_non_none([r["gross_margin_pct"] for r in data])
    avg_np     = _avg_non_none([r["net_profit"]       for r in data])
    avg_fc     = _avg_non_none([r["food_cost_pct"]    for r in data])
    avg_lc     = _avg_non_none([r["labor_cost_pct"]   for r in data])
    total_el   = sum(r["electricity"] or 0  for r in data)
    total_rent = sum(r["rent"]        or 0  for r in data)
    total_mkt  = sum(r["marketing"]   or 0  for r in data)
    total_pkg  = sum(r["packaging"]   or 0  for r in data)
    total_comm = sum(r["commission"]  or 0  for r in data)

    dates = sorted(set(r["date"] for r in data))
    daily = [{"date": d, "revenue": sum(r["daily_revenue"] for r in data if r["date"] == d)}
             for d in dates[-14:]]

    cost_breakdown = [
        {"label": "Electricity",  "value": round(total_el,   2)},
        {"label": "Rent",         "value": round(total_rent, 2)},
        {"label": "Marketing",    "value": round(total_mkt,  2)},
        {"label": "Packaging",    "value": round(total_pkg,  2)},
        {"label": "Commission",   "value": round(total_comm, 2)},
    ]

    return {
        "uploaded": True, "info": info,
        "summary": {
            "total_revenue":    round(total_rev, 2),
            "avg_gross_margin": round(avg_gm, 1) if avg_gm is not None else None,
            "avg_net_profit":   round(avg_np, 1) if avg_np is not None else None,
            "avg_food_cost":    round(avg_fc, 1) if avg_fc is not None else None,
            "avg_labor_cost":   round(avg_lc, 1) if avg_lc is not None else None,
            "records":          len(data),
            "days":             len(dates),
        },
        "daily_revenue": daily,
        "cost_breakdown": cost_breakdown,
        "recent": data[-20:],
    }


# ── POS Billing ────────────────────────────────────────────────────────────────

@router.post("/upload/pos")
async def upload_pos(file: UploadFile = File(...), mode: str = Query("replace"),
                     request: Request = None):
    _check_ext(file.filename)
    raw = await file.read()
    try:
        df = _read_file(raw, file.filename)
        records, info = _parse_pos(df, file.filename)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Parse error: {e}")

    if not records:
        raise HTTPException(status_code=422, detail="No valid rows found.")

    tid = _get_tenant_id(request)
    if mode == "append":
        existing, _ = data_store.get_pos_for_tenant(tid)
        if existing:
            existing_ids = {r["order_id"] for r in existing if r.get("order_id")}
            new_recs = [r for r in records if not r.get("order_id") or r["order_id"] not in existing_ids]
            duplicates = len(records) - len(new_recs)
            merged = existing + new_recs
            info["mode"] = "append"
            info["new_records"] = len(new_recs)
            info["duplicates_skipped"] = duplicates
            info["rows"] = len(merged)
            records = merged
        else:
            info["mode"] = "replace"
    else:
        info["mode"] = "replace"

    data_store.set_pos_for_tenant(tid, records, info)
    msg = (f"Added {info.get('new_records', len(records))} new POS records "
           f"({info.get('duplicates_skipped', 0)} duplicates skipped). "
           f"Total: {len(records):,} records."
           if mode == "append" else
           f"Loaded {len(records)} POS records.")
    return {"success": True, "message": msg, "info": info}


@router.delete("/upload/pos/clear")
def clear_pos(request: Request = None):
    tid = _get_tenant_id(request)
    data_store.clear_pos_for_tenant(tid)
    return {"success": True}


@router.get("/data/pos/summary")
def pos_summary(request: Request = None):
    tid = _get_tenant_id(request)
    data, info = data_store.get_pos_for_tenant(tid)
    if not data:
        return {"uploaded": False, "info": {}, "summary": {}}

    total_orders    = len(data)
    total_rev       = sum(r["bill_amount"] for r in data)
    avg_bill        = total_rev / max(total_orders, 1)
    total_discount  = sum(r["discount"] or 0 for r in data)
    total_gst       = sum(r["gst"]      or 0 for r in data)

    # Payment mode breakdown
    pay_agg: dict = {}
    for r in data:
        pm = r.get("payment_mode") or "Unknown"
        pay_agg[pm] = pay_agg.get(pm, 0) + 1

    # Platform breakdown
    plat_agg: dict = {}
    for r in data:
        p = r.get("platform") or "Unknown"
        plat_agg[p] = plat_agg.get(p, {"orders": 0, "revenue": 0.0})
        plat_agg[p]["orders"]  += 1
        plat_agg[p]["revenue"] += r["bill_amount"]

    # Peak hour breakdown
    hour_agg: dict = {}
    for r in data:
        label = r.get("peak_hour") or _hour_label(r.get("hour", 12))
        hour_agg[label] = hour_agg.get(label, 0) + 1

    # Repeat customer rate
    repeat_count = sum(1 for r in data
                       if str(r.get("repeat", "")).lower() in ("yes", "true", "1", "repeat", "returning"))
    repeat_pct   = round(repeat_count / max(total_orders, 1) * 100, 1)

    dates = sorted(set(r["date"] for r in data))

    return {
        "uploaded": True, "info": info,
        "summary": {
            "total_orders":   total_orders,
            "total_revenue":  round(total_rev, 2),
            "avg_bill":       round(avg_bill, 2),
            "total_discount": round(total_discount, 2),
            "total_gst":      round(total_gst, 2),
            "repeat_pct":     repeat_pct,
            "days":           len(dates),
        },
        "payment_breakdown": [{"mode": k, "count": v} for k, v in pay_agg.items()],
        "platform_breakdown": [{"platform": k, **v} for k, v in plat_agg.items()],
        "peak_hours": [{"slot": k, "orders": v} for k, v in sorted(hour_agg.items())],
        "recent": data[-20:],
    }


# ── Customer ───────────────────────────────────────────────────────────────────

@router.post("/upload/customer")
async def upload_customer(file: UploadFile = File(...), mode: str = Query("replace"),
                          request: Request = None):
    _check_ext(file.filename)
    raw = await file.read()
    try:
        df = _read_file(raw, file.filename)
        records, info = _parse_customer(df, file.filename)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Parse error: {e}")

    if not records:
        raise HTTPException(status_code=422, detail="No valid rows found.")

    tid = _get_tenant_id(request)
    if mode == "append":
        existing, _ = data_store.get_customer_for_tenant(tid)
        if existing:
            phone_map = {r["phone"]: i for i, r in enumerate(existing) if r.get("phone")}
            merged = list(existing)
            new_count, updated_count = 0, 0
            for r in records:
                phone = r.get("phone", "")
                if phone and phone in phone_map:
                    merged[phone_map[phone]] = r
                    updated_count += 1
                else:
                    merged.append(r)
                    new_count += 1
            info["mode"] = "append"
            info["new_records"] = new_count
            info["updated_records"] = updated_count
            info["rows"] = len(merged)
            info["total_customers"] = len(merged)
            records = merged
        else:
            info["mode"] = "replace"
    else:
        info["mode"] = "replace"

    data_store.set_customer_for_tenant(tid, records, info)
    msg = (f"Added {info.get('new_records', 0)} new + updated {info.get('updated_records', 0)} existing customers. "
           f"Total: {len(records):,} customers."
           if mode == "append" else
           f"Loaded {len(records)} customer records.")
    return {"success": True, "message": msg, "info": info}


@router.delete("/upload/customer/clear")
def clear_customer(request: Request = None):
    tid = _get_tenant_id(request)
    data_store.clear_customer_for_tenant(tid)
    return {"success": True}


@router.get("/data/customer/summary")
def customer_summary(request: Request = None):
    tid = _get_tenant_id(request)
    data, info = data_store.get_customer_for_tenant(tid)
    if not data:
        return {"uploaded": False, "info": {}, "summary": {}}

    # Text-to-numeric map for visit_frequency (visits per month)
    _FREQ_MAP = {
        "daily": 30, "everyday": 30,
        "weekly": 4, "every week": 4,
        "fortnightly": 2, "bi-weekly": 2, "biweekly": 2,
        "monthly": 1, "once a month": 1,
        "quarterly": 0.33, "rarely": 0.2, "occasional": 0.2, "sometimes": 0.5,
    }

    def _freq_to_num(v: str) -> float:
        key = str(v).lower().strip()
        for k, n in _FREQ_MAP.items():
            if k in key:
                return n
        try:
            return float(key)
        except Exception:
            return 0.0

    total      = len(data)
    avg_aov    = _avg_non_none([r["avg_order_value"] for r in data])
    avg_pts    = sum(r["loyalty_points"]  for r in data) / max(total, 1)

    # Frequency distribution (exact labels)
    freq_dist: dict = {}
    for r in data:
        f = r.get("visit_frequency") or "Not specified"
        freq_dist[f] = freq_dist.get(f, 0) + 1

    # Most common visit frequency label
    top_freq = max(freq_dist, key=freq_dist.get) if freq_dist else ""

    # Numeric average (visits/month) for summary widget
    freq_nums = [_freq_to_num(r.get("visit_frequency", "")) for r in data]
    avg_visits = sum(freq_nums) / max(total, 1)

    # Platform source
    src_agg: dict = {}
    for r in data:
        s = r.get("platform_source") or "Unknown"
        src_agg[s] = src_agg.get(s, 0) + 1

    # Gender
    gender_agg: dict = {}
    for r in data:
        g = r.get("gender") or "Not specified"
        gender_agg[g] = gender_agg.get(g, 0) + 1

    # Age group
    age_agg: dict = {}
    for r in data:
        a = r.get("age_group") or "Not specified"
        age_agg[a] = age_agg.get(a, 0) + 1

    # Feedback sentiment
    pos_fb  = sum(1 for r in data if _sentiment_is_positive(r.get("feedback", "")))
    neg_fb  = sum(1 for r in data if _sentiment_is_negative(r.get("feedback", "")))
    neut_fb = total - pos_fb - neg_fb

    return {
        "uploaded": True, "info": info,
        "summary": {
            "total_customers":  total,
            "avg_order_value":  round(avg_aov, 2) if avg_aov is not None else None,
            "avg_visit_freq":   round(avg_visits, 1),
            "top_visit_freq":   top_freq,
            "avg_loyalty_pts":  round(avg_pts, 1),
        },
        "visit_freq_dist":  [{"frequency": k, "count": v} for k, v in sorted(freq_dist.items(), key=lambda x: -x[1])],
        "platform_source":  [{"source": k, "count": v} for k, v in src_agg.items()],
        "gender_dist":      [{"gender": k, "count": v} for k, v in gender_agg.items()],
        "age_dist":         [{"age_group": k, "count": v} for k, v in age_agg.items()],
        "feedback_sentiment": [
            {"label": "Positive", "count": pos_fb},
            {"label": "Neutral",  "count": neut_fb},
            {"label": "Negative", "count": neg_fb},
        ],
        "recent": data[-20:],
    }


# ── Reviews / Sentiment ────────────────────────────────────────────────────────

@router.post("/upload/reviews")
async def upload_reviews(file: UploadFile = File(...)):
    _check_ext(file.filename)
    raw = await file.read()
    try:
        df = _read_file(raw, file.filename)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not read file: {e}")

    df.columns = [str(c).strip() for c in df.columns]

    # Flexible column normalisation — maps whatever the file uses to canonical names
    REVIEW_COL_ALIASES = {
        "Review_Text":      ["review_text", "review text", "review", "comment", "feedback text",
                             "customer review", "review content", "text", "comments"],
        "Sentiment_Label":  ["sentiment_label", "sentiment label", "sentiment", "label",
                             "category", "polarity", "tone"],
        "Rating":           ["rating", "star rating", "stars", "score", "review rating"],
        "Review_ID":        ["review_id", "review id", "id", "review no", "review number"],
        "Source":           ["source", "platform", "channel", "platform source", "review source"],
        "Review_Date":      ["review_date", "review date", "date", "review_date", "posted date",
                             "created at", "date of review"],
        "Cafe_Location":    ["cafe_location", "cafe location", "location", "branch", "outlet"],
        "Visit_Type":       ["visit_type", "visit type", "type of visit", "visit mode",
                             "dine in", "delivery"],
    }

    rename_map = {}
    for canonical, aliases in REVIEW_COL_ALIASES.items():
        if canonical in df.columns:
            continue  # already correct
        for col in df.columns:
            if _norm(col) in [_norm(a) for a in aliases]:
                rename_map[col] = canonical
                break
    if rename_map:
        df = df.rename(columns=rename_map)

    if "Review_Text" not in df.columns:
        raise HTTPException(
            status_code=422,
            detail="File must contain a 'Review_Text' (or similar) column. "
                   f"Found columns: {list(df.columns)}"
        )

    from sentiment_engine import get_engine
    engine = get_engine()
    try:
        records, stats = engine.process_dataframe(df, file.filename)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Sentiment processing failed: {e}")

    if not records:
        raise HTTPException(status_code=422, detail="No valid review rows found.")

    data_store._review_data = records
    data_store._review_info = engine.info
    engine.save_state()
    data_store.save_dataset("reviews_meta", [], engine.info)   # persist info; full records in engine state

    return {
        "success": True,
        "message": f"Analysed {len(records)} reviews using TF-IDF + Linear SVM model.",
        "info": engine.info,
        "stats_preview": {
            "positive":          stats.get("positive", 0),
            "neutral":           stats.get("neutral",  0),
            "negative":          stats.get("negative", 0),
            "satisfaction_score": stats.get("satisfaction_score", 0),
        },
    }


@router.delete("/upload/reviews/clear")
def clear_reviews():
    from sentiment_engine import get_engine
    engine = get_engine()
    engine._records = []
    engine._stats   = {}
    engine._info    = {}
    engine.clear_state()
    data_store._review_data = []
    data_store._review_info = {}
    data_store.clear_dataset("reviews_meta")
    return {"success": True}


@router.get("/data/reviews/summary")
def reviews_summary():
    from sentiment_engine import get_engine
    engine = get_engine()
    if not engine.has_data:
        return {"uploaded": False, "info": {}, "summary": {}, "stats": {}}
    return {
        "uploaded": True,
        "info":     engine.info,
        "stats":    engine.stats,
        "recent":   engine.records[-10:],
    }


# ─────────────────────────────────────────────
# 5. MENU DATA
# ─────────────────────────────────────────────

MENU_ALIASES = {
    "category": ["category", "cat", "item category", "menu category", "food category"],
    "item":     ["item name", "item", "product", "dish", "menu item", "product name", "description"],
    # "cost" deliberately excluded — cost ≠ selling price; "rate" kept as it maps to selling price
    "base_price": ["base price", "base price (rs)", "base price(rs)", "base_price",
                   "base rate", "mrp", "selling price", "rate", "price"],
    "season":   ["season", "season code", "availability season", "seasonal"],
    "dayparts": ["available dayparts", "dayparts", "applicable dayparts", "time slots",
                 "daypart", "meal period"],
    # "type" removed — too generic and conflicts with "category"; use explicit veg/non-veg phrases
    "veg":      ["veg / non-veg", "veg/non-veg", "veg non veg", "veg or non-veg",
                 "food type", "veg status", "vegetarian"],
    "notes":    ["notes", "note", "remarks", "comments", "special notes"],
    "sku":      ["sku", "sku id", "sku_id", "item code", "product code", "code"],
}


def _parse_menu(raw: bytes, filename: str) -> tuple[list, dict]:
    import openpyxl
    import io as _io

    # Try to open as Excel; fall back to CSV
    if filename.lower().endswith(".csv"):
        df = pd.read_csv(_io.BytesIO(raw))
    else:
        # Try "Master Menu" sheet first, then first sheet
        wb = openpyxl.load_workbook(_io.BytesIO(raw), data_only=True)
        sheet_name = None
        for name in wb.sheetnames:
            if "master" in name.lower() or "menu" in name.lower():
                sheet_name = name
                break
        sheet_name = sheet_name or wb.sheetnames[0]
        df = pd.read_excel(_io.BytesIO(raw), sheet_name=sheet_name, engine="openpyxl")

    # Drop fully-empty rows and reset
    df.dropna(how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)

    # The sheet often has a merged title row at top — drop rows until we find the header
    # Strategy: find the row whose cells best match our known column names
    header_row = 0
    best_score = 0
    all_known = [alias for aliases in MENU_ALIASES.values() for alias in aliases]
    for i in range(min(5, len(df))):
        row_vals = [str(v).lower().strip() for v in df.iloc[i] if pd.notna(v)]
        score = sum(1 for v in row_vals if v in all_known)
        if score > best_score:
            best_score = score
            header_row = i

    if header_row > 0:
        df.columns = [str(v).strip() for v in df.iloc[header_row]]
        df = df.iloc[header_row + 1:].reset_index(drop=True)
    else:
        df.columns = [str(c).strip() for c in df.columns]

    cat_col   = _detect(df.columns.tolist(), MENU_ALIASES, "category")
    item_col  = _detect(df.columns.tolist(), MENU_ALIASES, "item")
    price_col = _detect(df.columns.tolist(), MENU_ALIASES, "base_price")
    season_col= _detect(df.columns.tolist(), MENU_ALIASES, "season")
    day_col   = _detect(df.columns.tolist(), MENU_ALIASES, "dayparts")
    veg_col   = _detect(df.columns.tolist(), MENU_ALIASES, "veg")
    notes_col = _detect(df.columns.tolist(), MENU_ALIASES, "notes")
    sku_col   = _detect(df.columns.tolist(), MENU_ALIASES, "sku")

    if not item_col:
        raise ValueError(
            f"Could not find an 'Item' / 'Item Name' column. "
            f"Found columns: {list(df.columns)}"
        )

    records, skipped = [], 0
    categories: dict = {}
    for _, row in df.iterrows():
        try:
            item = _safe_str(row[item_col]) if item_col else ""
            if not item or item.lower() in ("nan", "none", ""):
                skipped += 1
                continue

            cat   = _safe_str(row[cat_col],   "Uncategorized") if cat_col   else "Uncategorized"
            price = _safe_float(row[price_col], 0.0)            if price_col else 0.0
            season= _safe_str(row[season_col], "YR")            if season_col else "YR"
            days  = _safe_str(row[day_col],    "")              if day_col    else ""
            veg   = _safe_str(row[veg_col],    "")              if veg_col    else ""
            notes = _safe_str(row[notes_col],  "")              if notes_col  else ""
            sku   = _safe_str(row[sku_col],    "")              if sku_col    else ""

            records.append({
                "sku":      sku,
                "category": cat,
                "item":     item,
                "base_price": price,
                "season":   season,
                "dayparts": days,
                "veg":      veg,
                "notes":    notes,
            })
            categories[cat] = categories.get(cat, 0) + 1
        except Exception:
            skipped += 1

    prices = [r["base_price"] for r in records if r["base_price"] > 0]

    detected = {k: v for k, v in {
        "category": cat_col, "item": item_col, "base_price": price_col,
        "season": season_col, "dayparts": day_col, "veg": veg_col,
        "notes": notes_col, "sku": sku_col,
    }.items() if v}

    seasons_found = sorted(set(r["season"] for r in records if r["season"]))

    info = {
        "filename":  filename,
        "rows":      len(records),
        "skipped":   skipped,
        "columns_detected": detected,
        "total_skus":    len(records),
        "total_categories": len(categories),
        "price_range": {
            "min": round(min(prices), 2) if prices else 0,
            "max": round(max(prices), 2) if prices else 0,
        },
        "seasons": seasons_found,
        "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    return records, info


@router.post("/upload/menu")
async def upload_menu(file: UploadFile = File(...), request: Request = None):
    if not any(file.filename.lower().endswith(ext) for ext in (".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, or .csv files are supported.")
    raw = await file.read()
    try:
        records, info = _parse_menu(raw, file.filename)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Parse error: {e}")

    if not records:
        raise HTTPException(status_code=422, detail="No valid menu rows found.")

    tid = _get_tenant_id(request)
    data_store.set_menu_for_tenant(tid, records, info)
    return {"success": True, "message": f"Loaded {len(records)} menu items.", "info": info}


@router.delete("/upload/menu/clear")
def clear_menu(request: Request = None):
    tid = _get_tenant_id(request)
    data_store.clear_menu_for_tenant(tid)
    return {"success": True}


@router.get("/data/menu/summary")
def menu_summary(request: Request = None):
    tid = _get_tenant_id(request)
    data, info = data_store.get_menu_for_tenant(tid)
    if not data:
        return {"uploaded": False, "info": {}, "summary": {}}

    # Category breakdown
    cat_agg: dict = {}
    for r in data:
        c = r["category"]
        cat_agg[c] = cat_agg.get(c, 0) + 1

    # Veg/Non-Veg split
    veg_agg: dict = {}
    for r in data:
        v = r["veg"] or "Not specified"
        veg_agg[v] = veg_agg.get(v, 0) + 1

    # Season breakdown
    season_agg: dict = {}
    for r in data:
        s = r["season"] or "YR"
        season_agg[s] = season_agg.get(s, 0) + 1

    prices = [r["base_price"] for r in data if r["base_price"] > 0]

    return {
        "uploaded": True,
        "info": info,
        "summary": {
            "total_skus":        len(data),
            "total_categories":  len(cat_agg),
            "price_min":         round(min(prices), 0) if prices else 0,
            "price_max":         round(max(prices), 0) if prices else 0,
            "price_avg":         round(sum(prices) / len(prices), 0) if prices else 0,
            "seasonal_items":    sum(v for k, v in season_agg.items() if k != "YR"),
            "year_round_items":  season_agg.get("YR", 0),
        },
        "category_breakdown": [{"category": k, "count": v} for k, v in sorted(cat_agg.items(), key=lambda x: -x[1])],
        "veg_breakdown":      [{"type": k, "count": v} for k, v in veg_agg.items()],
        "season_breakdown":   [{"season": k, "count": v} for k, v in season_agg.items()],
        "recent": data[:20],
    }


# ── Paginated records viewer ───────────────────────────────────────────────────

@router.get("/data/{dtype}/records")
def get_records(dtype: str, page: int = 1, per_page: int = 50, search: str = "",
                request: Request = None):
    """Return a paginated, optionally searched slice of any uploaded dataset."""
    from sentiment_engine import get_engine
    tid = _get_tenant_id(request)
    fin_data, _  = data_store.get_financial_for_tenant(tid)
    pos_data, _  = data_store.get_pos_for_tenant(tid)
    cust_data, _ = data_store.get_customer_for_tenant(tid)
    menu_data, _ = data_store.get_menu_for_tenant(tid)

    dtype_map = {
        "financial": fin_data,
        "pos":       pos_data,
        "customer":  cust_data,
        "menu":      menu_data,
    }
    if dtype == "reviews":
        data = get_engine().records
    elif dtype in dtype_map:
        data = dtype_map[dtype]
    else:
        raise HTTPException(status_code=400, detail=f"Unknown dataset type: {dtype}")

    if not data:
        return {"total": 0, "page": page, "per_page": per_page, "pages": 0, "records": []}

    # Search: filter rows where any string value contains the search term
    if search:
        q = search.lower()
        filtered = [r for r in data if any(q in str(v).lower() for v in r.values())]
    else:
        filtered = data

    total = len(filtered)
    pages = max(1, (total + per_page - 1) // per_page)
    page  = max(1, min(page, pages))
    start = (page - 1) * per_page
    end   = start + per_page

    return {
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "pages":    pages,
        "records":  filtered[start:end],
    }


# ── Aggregate status ───────────────────────────────────────────────────────────

@router.get("/upload/status/all")
def upload_status_all(request: Request = None):
    from sentiment_engine import get_engine
    engine = get_engine()
    tid = _get_tenant_id(request)
    fin_data,  fin_info  = data_store.get_financial_for_tenant(tid)
    pos_data,  pos_info  = data_store.get_pos_for_tenant(tid)
    cust_data, cust_info = data_store.get_customer_for_tenant(tid)
    menu_data, menu_info = data_store.get_menu_for_tenant(tid)
    return {
        "financial": {
            "uploaded": bool(fin_data),
            "info":     fin_info,
        },
        "pos": {
            "uploaded": bool(pos_data),
            "info":     pos_info,
        },
        "customer": {
            "uploaded": bool(cust_data),
            "info":     cust_info,
        },
        "reviews": {
            "uploaded": engine.has_data,
            "info":     engine.info,
        },
        "menu": {
            "uploaded": bool(menu_data),
            "info":     menu_info,
        },
    }


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def _avg_non_none(values: list) -> Optional[float]:
    vals = [v for v in values if v is not None]
    return sum(vals) / len(vals) if vals else None


def _sentiment_is_positive(text: str) -> bool:
    text = str(text).lower()
    try:
        score = float(text)
        return score >= 4
    except ValueError:
        pass
    pos_words = {"good", "great", "excellent", "loved", "amazing", "happy",
                 "satisfied", "positive", "superb", "fantastic", "awesome"}
    return any(w in text for w in pos_words)


def _sentiment_is_negative(text: str) -> bool:
    text = str(text).lower()
    try:
        score = float(text)
        return score <= 2
    except ValueError:
        pass
    neg_words = {"bad", "poor", "terrible", "worst", "unhappy", "disappointed",
                 "negative", "horrible", "awful", "disgust", "rude", "slow"}
    return any(w in text for w in neg_words)
